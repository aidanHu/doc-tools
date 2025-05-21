import os
import pandas as pd
from openpyxl import Workbook, load_workbook
import time
from openpyxl.utils import get_column_letter

def collect_txt_files(folder_path):
    """收集指定文件夹及其子文件夹中的所有txt文件"""
    print(f"开始扫描文件夹: {folder_path}")
    txt_files = []
    
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith('.txt'):
                full_path = os.path.join(root, file)
                txt_files.append(full_path)
                
    print(f"扫描完成，共找到 {len(txt_files)} 个txt文件")
    return txt_files

def read_txt_content(file_path):
    """读取txt文件的内容"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        return content
    except UnicodeDecodeError:
        # 如果UTF-8解码失败，尝试使用其他编码
        try:
            with open(file_path, 'r', encoding='gbk') as file:
                content = file.read()
            return content
        except Exception as e:
            return f"无法读取文件: {str(e)}"
    except Exception as e:
        return f"读取文件时出错: {str(e)}"

def create_excel_if_not_exists(output_excel):
    """如果Excel文件不存在，则创建一个新的带有标题行的Excel文件"""
    if not os.path.exists(output_excel):
        wb = Workbook()
        ws = wb.active
        ws.title = "TXT文件内容"
        
        # 添加标题行
        headers = ['文件名', '相对路径', '文件内容']
        ws.append(headers)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30  # 文件名
        ws.column_dimensions['B'].width = 50  # 相对路径
        ws.column_dimensions['C'].width = 100  # 文件内容
        
        wb.save(output_excel)
        print(f"创建了新的Excel文件: {output_excel}")

def process_and_save_files(txt_files, output_excel, folder_path):
    """逐个处理txt文件并保存到Excel"""
    print("开始处理文件内容...")
    start_time = time.time()
    total_files = len(txt_files)
    
    # 创建Excel文件（如果不存在）
    create_excel_if_not_exists(output_excel)
    
    # 记录已处理的文件，避免重复处理
    processed_files = set()
    
    try:
        # 检查是否有已处理的文件
        if os.path.exists(output_excel):
            wb = load_workbook(output_excel)
            ws = wb.active
            
            # 从第2行开始（跳过标题行）
            for row in range(2, ws.max_row + 1):
                file_path = ws.cell(row=row, column=2).value
                if file_path:
                    processed_files.add(file_path)
            
            print(f"检测到已处理 {len(processed_files)} 个文件")
    except Exception as e:
        print(f"读取已有Excel文件时出错: {str(e)}")
        print("将创建新的Excel文件")
        create_excel_if_not_exists(output_excel)
    
    # 处理文件
    success_count = 0
    error_count = 0
    
    for index, file_path in enumerate(txt_files, 1):
        relative_path = os.path.relpath(file_path, start=folder_path)
        
        # 如果文件已经处理过，则跳过
        if relative_path in processed_files:
            print(f"跳过已处理的文件 ({index}/{total_files}): {os.path.basename(file_path)}")
            continue
            
        print(f"正在处理文件 ({index}/{total_files}): {os.path.basename(file_path)}")
        
        try:
            # 读取文件内容
            content = read_txt_content(file_path)
            file_name = os.path.basename(file_path)
            
            # 加载当前Excel
            wb = load_workbook(output_excel)
            ws = wb.active
            
            # 添加新行
            new_row = [file_name, relative_path, content]
            ws.append(new_row)
            
            # 保存Excel
            wb.save(output_excel)
            success_count += 1
            print(f"  ✓ 成功处理并保存: {file_name}")
            
        except Exception as e:
            error_count += 1
            print(f"  ✗ 处理文件时出错: {file_path}")
            print(f"    错误信息: {str(e)}")
    
    elapsed_time = time.time() - start_time
    print("\n处理完成!")
    print(f"成功处理: {success_count} 个文件")
    print(f"处理失败: {error_count} 个文件")
    print(f"总耗时: {elapsed_time:.2f} 秒")

if __name__ == "__main__":
    # 设置文件夹路径和输出Excel文件名
    folder_path = input("请输入要扫描的文件夹路径: ")
    output_excel = input("请输入输出Excel文件的路径(例如 output.xlsx): ")
    
    # 显示完整路径
    folder_path = os.path.expanduser(folder_path)  # 展开 ~ 符号
    output_excel = os.path.expanduser(output_excel)
    # folder_path = input("/Users/aidan/Desktop/女性向文案号")
    # output_excel = input("/Users/aidan/Desktop/output.xlsx")
    
    print(f"使用的文件夹路径: {folder_path}")
    print(f"输出Excel文件将保存为: {output_excel}")
    
    # 收集所有txt文件
    txt_files = collect_txt_files(folder_path)
    
    # 处理并保存文件
    if txt_files:
        process_and_save_files(txt_files, output_excel, folder_path)
    else:
        print("未找到任何txt文件，程序结束")


